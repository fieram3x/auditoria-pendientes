from datetime import datetime
from io import BytesIO

import pandas as pd

from constants import CLOSED_STATUS
from ui_utils import normalize_text, safe_date


def _priority_fill(priority):
    priority = normalize_text(priority)
    return {
        "Crítica": "FEE2E2",
        "Critica": "FEE2E2",
        "Alta": "FFEDD5",
        "Media": "FEF9C3",
        "Baja": "DCFCE7",
    }.get(priority, "FFFFFF")


def _status_fill(status):
    status = normalize_text(status)
    if status in CLOSED_STATUS:
        return "DCFCE7"
    return {
        "Pendiente": "F1F5F9",
        "En proceso": "DBEAFE",
        "En espera de respuesta": "FEF3C7",
        "Escalado": "EDE9FE",
    }.get(status, "FFFFFF")


def filtered_excel_bytes(df, filters=None, title="Incidencias filtradas"):
    output = BytesIO()
    table_df = df.copy().fillna("")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        table_df.to_excel(writer, index=False, sheet_name="Incidencias")
        wb = writer.book
        ws = writer.sheets["Incidencias"]

        from openpyxl.styles import Alignment, Font, PatternFill

        header_fill = PatternFill("solid", fgColor="DBEAFE")
        header_font = Font(bold=True, color="0F172A")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        priority_col = None
        status_col = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Prioridad":
                priority_col = idx
            if cell.value == "Estatus":
                status_col = idx

        for row in range(2, ws.max_row + 1):
            if priority_col:
                ws.cell(row, priority_col).fill = PatternFill("solid", fgColor=_priority_fill(ws.cell(row, priority_col).value))
            if status_col:
                ws.cell(row, status_col).fill = PatternFill("solid", fgColor=_status_fill(ws.cell(row, status_col).value))
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)

        for col_cells in ws.columns:
            values = [str(cell.value or "") for cell in col_cells]
            width = min(max(len(value) for value in values) + 2, 42)
            ws.column_dimensions[col_cells[0].column_letter].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        info = wb.create_sheet("Filtros")
        info["A1"] = title
        info["A1"].font = Font(bold=True, size=14)
        info["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %I:%M %p')}"
        row_num = 4
        for item in filters or []:
            info.cell(row_num, 1).value = item
            row_num += 1
        info.column_dimensions["A"].width = 80

    return output.getvalue()


def dashboard_pdf_bytes(df, kpis, filters=None):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception as exc:
        raise RuntimeError("Instala reportlab para generar el PDF ejecutivo.") from exc

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=26,
        leftMargin=26,
        topMargin=24,
        bottomMargin=24,
        title="Dashboard Auditoría Pendientes",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "DashboardTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#64748b"),
        spaceAfter=10,
    )
    section_style = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        textColor=colors.HexColor("#0f172a"),
        spaceBefore=8,
        spaceAfter=6,
    )
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=7.2, leading=9, textColor=colors.HexColor("#334155"))

    story = [
        Paragraph("Auditoría Pendientes", title_style),
        Paragraph(f"Reporte ejecutivo generado el {datetime.now().strftime('%d/%m/%Y %I:%M %p')}", subtitle_style),
    ]

    filter_rows = [[Paragraph("<b>Filtros aplicados</b>", small)]]
    for item in filters or ["Sin filtros"]:
        filter_rows.append([Paragraph(str(item), small)])
    filter_table = Table(filter_rows, colWidths=[10.2 * inch])
    filter_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eff6ff")),
        ("BOX", (0, 0), (-1, -1), .6, colors.HexColor("#dbe7f5")),
        ("INNERGRID", (0, 0), (-1, -1), .25, colors.HexColor("#e5edf7")),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.extend([filter_table, Spacer(1, 8)])

    kpi_cells = []
    for label, value, sub in kpis[:5]:
        kpi_cells.append(Paragraph(f"<b>{label}</b><br/><font size='14'><b>{value}</b></font><br/><font color='#64748b'>{sub}</font>", small))
    kpi_table = Table([kpi_cells], colWidths=[2.0 * inch] * len(kpi_cells))
    kpi_table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), .6, colors.HexColor("#dbe7f5")),
        ("INNERGRID", (0, 0), (-1, -1), .25, colors.HexColor("#e5edf7")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.extend([Paragraph("Indicadores", section_style), kpi_table, Spacer(1, 8)])

    story.append(Paragraph("Resumen ejecutivo", section_style))
    total = len(df)
    open_count = len(df[~df["Estatus"].astype(str).isin(CLOSED_STATUS)]) if "Estatus" in df.columns else 0
    critical = len(df[df["Prioridad"].astype(str).isin(["Crítica", "Critica"])]) if "Prioridad" in df.columns else 0
    story.append(Paragraph(f"El reporte contiene {total} incidencia(s), con {open_count} abierta(s) y {critical} crítica(s).", small))
    story.append(Spacer(1, 8))

    story.append(Paragraph("Tabla resumida", section_style))
    cols = [
        "ID",
        "Fecha Creación",
        "Hotel",
        "Departamento",
        "Prioridad",
        "Estatus",
        "Responsable",
        "SLA",
        "Fecha Compromiso",
        "Descripción",
    ]
    table_df = df[[col for col in cols if col in df.columns]].copy().head(28)
    for col in ["Fecha Creación", "Fecha Compromiso"]:
        if col in table_df.columns:
            table_df[col] = table_df[col].apply(safe_date)

    if table_df.empty:
        story.append(Paragraph("No hay incidencias con los filtros seleccionados.", small))
    else:
        rows = [[Paragraph(f"<b>{col}</b>", small) for col in table_df.columns]]
        for _, row in table_df.iterrows():
            rows.append([Paragraph(str(row.get(col, ""))[:450], small) for col in table_df.columns])
        widths = [0.82 * inch, .82 * inch, .9 * inch, .95 * inch, .68 * inch, .78 * inch, .95 * inch, .85 * inch, .82 * inch, 2.35 * inch][:len(table_df.columns)]
        detail_table = Table(rows, colWidths=widths, repeatRows=1)
        detail_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8fbff")),
            ("BOX", (0, 0), (-1, -1), .6, colors.HexColor("#dbe7f5")),
            ("INNERGRID", (0, 0), (-1, -1), .25, colors.HexColor("#e5edf7")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(detail_table)

    doc.build(story)
    return output.getvalue()
